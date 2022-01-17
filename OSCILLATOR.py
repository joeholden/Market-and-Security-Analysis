import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
import pyexcel as p
import re
import requests
import os
from datetime import datetime
import numpy as np
import statistics

HEADERS = {
    'a': 'date',
    'b': 'NYSE Advancing Issues',
    'c': 'NYSE Declining issues',
    'd': 'NYSE up volume',
    'e': 'NYSE down volume',
    'f': 'DJIA Close',
    'g': 'NYSE A-D',
    'h': '10% Trend A-D',
    'i': '5% Trend A-D',
    'j': 'McC A-D Osc',
    'k': 'McC Summation Index',
    'l': 'Osc unchanged tomorrow',
    'm': 'Osc to zero tomorrow',
    'n': 'NYSE UV-DV',
    'o': '10% Trend UV-DV',
    'p': '5% Trend UV-DV',
    'q': 'McC UV-DV Osc',
    'r': 'McC Vol Summation Index',
    's': 'UV-DV for unchanged osc tomorrow',
    't': 'UV-DV for osc to zero tomorrow',
    'u': '10% trend',
    'v': '5% trend',
    'w': 'price oscillator',
    'x': 'price for unchanged oscillator',
}


def download_mclellan_website_data():
    url = 'https://www.mcoscillator.com/market_breadth_data/'
    s = Service(ChromeDriverManager().install())
    op = webdriver.ChromeOptions()
    # op.add_argument('--headless')
    # op.add_argument('--disable-gpu')
    driver = webdriver.Chrome(service=s, options=op)

    driver.get(url)
    download_link = driver.find_element(by=By.XPATH, value='//*[@id="data_table"]/a[1]/img')
    download_link.click()

    time.sleep(1)


def clean_excel_data():
    """Takes the xls file that was downloaded from download_mclellan_website_data, converts it to .xlsx
    and then uses openpyxl to process it to a cleaned dataframe.
    """
    global first_row_to_keep, first_row_to_delete, start_date, end_date

    # Change format to.xlsx instead of .xls so openpyxl can use it
    p.save_book_as(file_name='C:/Users/joema/Downloads/OSC-DATA.xls',
                   dest_file_name='C:/Users/joema/Downloads/OSC-DATA.xlsx')
    wb = openpyxl.load_workbook('C:/Users/joema/Downloads/OSC-DATA.xlsx')
    ws = wb['A']

    # Get the correct bounds of the dataframe
    # the group(0) notation just accesses the text of the match object
    # Checks first 30 rows for a date-like object

    for cell_row in range(1, 30):
        c = ws[f'A{cell_row}']
        cell_contents = str(c.value).split(" ")[0]
        try:
            re.match(r"^([\d]{4})-([\d]+)-([\d]+)$", cell_contents).group(0)
            first_row_to_keep = cell_row
            start_date = ws[f'A{cell_row}'].value
            break
        except AttributeError:
            pass
    # first row to keep: most likely this will be cell A9 unless the format changes

    # Check for first row to delete
    # if cell.value: just checks if the cell exists. In openpyxl cells are created as they are populated
    # They do not exist otherwise

    for cell_row in range(first_row_to_keep, 500):
        if not ws[f'B{cell_row}'].value:
            first_row_to_delete = cell_row
            end_date = ws[f'A{cell_row - 1}'].value
            break

    # delete rows to clean
    ws.delete_rows(1, first_row_to_keep - 1)
    ws.delete_rows(first_row_to_delete - first_row_to_keep + 1, 500)
    wb.save('excel files/cleaned.xlsx')


def get_advance_decline_data(timeframe):
    """gets a plot of the advance decline line (10 day sma)"""
    data = pd.read_excel('excel files/cleaned.xlsx', engine='openpyxl', names=HEADERS.values(), )
    advance_decline = data['NYSE A-D']
    date = data['date']
    data['advance_decline_line'] = [j + (j - 1) for j in advance_decline]

    data[f'{timeframe} day sma'] = data['advance_decline_line'].rolling(window=timeframe).mean()

    if timeframe == 10:
        sma = data[f'{timeframe} day sma']
    elif timeframe == 30:
        sma = data[f'{timeframe} day sma']
    else:
        raise ValueError

    return date, advance_decline, sma


def get_volume_data():
    """Gets the 30 day rolling average of % up volume NYSE"""
    data = pd.read_excel('excel files/cleaned.xlsx', engine='openpyxl', names=HEADERS.values())
    up_vol = data['NYSE up volume']
    down_vol = data['NYSE down volume']
    percent_up_volume = round((up_vol / (up_vol + down_vol)) * 100, 0)

    data['30 day sma'] = percent_up_volume.rolling(window=30).mean()

    date = data['date']
    sma30 = data['30 day sma']

    return date, percent_up_volume, sma30


def plot_mcclellan_index_and_reversal(benchmark_index):
    """
    Ratio-Adjusted Net Advances (RANA): (Advances - Declines)/(Advances + Declines)
    McClellan Oscillator: 19-day EMA of RANA - 39-day EMA of RANA
    The first EMA calculation is a simple average.
    Mcclellan Index is the sum of all previous day's values for the oscillator
    When the Mcclellan Summation Index requires +-4,000 issues to turn, we are oversold/overbought
    index_today + oscillator_tomrrow = index_tomorrow -> for a reversal, oscillator has to cross zero
    """
    data = pd.read_excel('excel files/cleaned.xlsx', engine='openpyxl', names=HEADERS.values())
    mcclellan_summation_index = data['McC Summation Index']
    mcclellan_oscillator = data['McC A-D Osc']
    oscillator_to_zero = data['Osc to zero tomorrow']

    spy_data = benchmark_index
    spy_data_dates = [datetime.strptime(i, "%Y-%m-%d") for i in list(spy_data.keys())]

    fig, ax = plt.subplots(2, figsize=(13, 10), gridspec_kw={'height_ratios': [3, 1]})

    ax0 = ax[0]
    ax1 = ax0.twinx()
    ax2 = ax[1]

    ax0.plot(d, mcclellan_summation_index, color='purple', label='McClellan Summation Index')
    ax1.plot(spy_data_dates, spy_data.values(), color='blue', label='SPY')

    ax0.tick_params(axis='y', labelcolor='purple')
    ax1.tick_params(axis='y', labelcolor='blue')

    h1, l1 = ax0.get_legend_handles_labels()
    h2, l2 = ax1.get_legend_handles_labels()
    ax0.legend(h1 + h2, l1 + l2, loc=0, facecolor='#bebebe', framealpha=0, fontsize=12)
    ax0.set_facecolor('#bebebe')

    plt.title(f'McClellan Summation Index vs. SPY\n', fontsize=20)

    ax2.bar(data["date"], oscillator_to_zero, color='purple')
    ax2.set_facecolor('#bebebe')
    ax2.set_title('A-D to Turn McClellan Summation Index', size=18)
    plt.figtext(0.91, 0.115, 'Overbought', color='red')
    plt.figtext(0.91, 0.25, 'Oversold', color='green')
    plt.sca(ax2)

    overbought_oversold_points = []
    plt.axhline(y=4000, linestyle='--', color='purple')
    plt.axhline(y=-4000, linestyle='--', color='purple')
    for i, v in enumerate(oscillator_to_zero):
        if v > 4000:
            overbought_oversold_points.append((i, v))
        if v < -4000:
            overbought_oversold_points.append((i, v))

    plt.sca(ax1)
    ax1.scatter([data['date'][i] for (i, v) in overbought_oversold_points],
                [spy_data[str(data['date'][i]).split(" ")[0]] for (i, v) in overbought_oversold_points], color='black')

    current_oscillator_to_zero = round(list(oscillator_to_zero)[-1], 0)
    plt.figtext(0.34, 0.89, f'Current A-D to Turn Index: {current_oscillator_to_zero}', color='purple', fontsize=20)

    plt.savefig('plots/mcclellan.png')


def get_spy():
    response = requests.get(f'https://financialmodelingprep.com/api/v3/historical-price-full/SPY?serietype=line'
                            f'&apikey={os.environ["APIKEY"]}')
    data = response.json()['historical']

    spy_range_data = {}
    for branch in data:
        date = branch['date']
        if start_date <= datetime.strptime(date, "%Y-%m-%d") <= end_date:
            spy_range_data[f'{branch["date"]}'] = branch['close']

    return spy_range_data


def plot_ad():
    spy_data_dates = [datetime.strptime(i, "%Y-%m-%d") for i in list(spy_data.keys())]

    fig, ax1 = plt.subplots(figsize=(18, 9))

    ax1.plot(d, sma_ad, color='purple', label='Oscillator')
    ax1.scatter(d, sma_ad, color='purple', s=7)
    ax1.plot(d, [0] * len(d), '--', color='white')

    ax2 = plt.twinx(ax1)
    ax2.plot(spy_data_dates, spy_data.values(), color='blue', label='SPY')
    ax2.scatter(spy_data_dates, spy_data.values(), color='blue', s=7)

    # ax2.spines['left'].set_color('purple')
    # ax2.spines['right'].set_color('blue')
    ax1.tick_params(axis='y', labelcolor='purple')
    ax2.tick_params(axis='y', labelcolor='blue')
    ax1.tick_params(axis='both', which='major', labelsize=17)
    ax2.tick_params(axis='both', which='major', labelsize=17)

    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, facecolor='#bebebe', framealpha=0, bbox_to_anchor=(1.01, 1), loc="upper left", fontsize=14)
    ax1.set_facecolor('#bebebe')

    plt.title(f'Advance Decline Oscillator ({timeframe}-Day) vs. SPY', fontsize=25)
    plt.tight_layout()
    plt.savefig('plots/ad.png')


def plot_vol():
    spy_data_dates = [datetime.strptime(i, "%Y-%m-%d") for i in list(spy_data.keys())]

    fig, ax1 = plt.subplots(figsize=(18, 9))

    ax1.plot(d, sma_v, color='purple', label='Oscillator')
    ax1.scatter(d, sma_v, color='purple', s=7)
    ax1.plot(d, [45] * len(d), '--', color='#128059', label='Near End of Selling')
    ax1.plot(d, [55] * len(d), '--', color='black', label='Near End of Buying')

    max_vol = np.nanmax(np.array(sma_v))
    min_vol = np.nanmin(np.array(sma_v))
    ax1.set_ylim(0.95 * min_vol, 1.05 * max_vol)

    ax2 = plt.twinx(ax1)
    ax2.plot(spy_data_dates, spy_data.values(), color='blue', label='SPY')
    ax2.scatter(spy_data_dates, spy_data.values(), color='blue', s=7)

    ax1.tick_params(axis='y', labelcolor='purple')
    ax2.tick_params(axis='y', labelcolor='blue')
    ax1.set_ylabel('Percent Up Volume NYSE', color='purple', fontsize=20)
    ax2.set_ylabel('SPY Price', color='blue', fontsize=20, rotation=270, labelpad=15)
    ax1.tick_params(axis='both', which='major', labelsize=17)
    ax2.tick_params(axis='both', which='major', labelsize=17)

    h1, l1 = ax1.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax1.legend(h1 + h2, l1 + l2, facecolor='#bebebe', framealpha=0, bbox_to_anchor=(1.01, 1), loc="upper left", fontsize=14)
    ax1.set_facecolor('#bebebe')

    plt.title(f'Volume Oscillator (30-Day) vs. SPY', fontsize=25)
    plt.tight_layout()
    plt.savefig('plots/vol.png')


def clean_up():
    downloaded_files = os.listdir('C:/Users/joema/Downloads/')
    for f in downloaded_files:
        if 'OSC-DATA' in f:
            os.remove(f'C:/Users/joema/Downloads/{f}')


timeframe = 10
download_mclellan_website_data()
clean_excel_data()
d, ad, sma_ad = get_advance_decline_data(timeframe=timeframe)
d_, up_v_percent, sma_v = get_volume_data()
spy_data = get_spy()
plot_mcclellan_index_and_reversal(spy)
plot_ad()
plot_vol()
clean_up()
