import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import datetime
from datetime import date, timedelta
import matplotlib.pyplot as plt
import matplotlib
import time

"""
This script scrapes CBOE website for total and equity put/call ratios and updates a central excel file.
This file will be used to create a put call oscillator plot
Set to run in Task Scheduler every day at 4pm under the name 'TEST'
"""

get_historical = False
get_today = True
plot_figure = True
# 'epc' or 'tpc' | equity or total
which_option_to_plot = 'tpc'

timeframe_to_plot_sma = 10
date_axis_plotting_interval = 15
# If you change get_historical to True, change start and end dates. Its non-inclusive at the end date
start_date = date(2022, 1, 21)
end_date = date(2022, 2, 2)


def download_put_call():
    # url = 'https://www.cboe.com/us/options/market_statistics/daily/'
    url = f'https://www.cboe.com/us/options/market_statistics/daily/?dt={single_date}'
    s = Service(ChromeDriverManager().install())
    op = webdriver.ChromeOptions()
    op.headless = True
    driver = webdriver.Chrome(service=s, options=op)

    driver.get(url)
    total_put_call = driver.find_element(by=By.XPATH,
                                         value='/html/body/main/section/div/div[2]/table/tbody/tr[1]/td[2]')
    equity_put_call = driver.find_element(by=By.XPATH,
                                          value='/html/body/main/section/div/div[2]/table/tbody/tr[4]/td[2]')
    return total_put_call.text, equity_put_call.text


def update_excel_sheet():
    headers = ['Day', 'Total Put/Call', 'Equity Put/Call']
    wb = load_workbook('C:/Users/joema/PycharmProjects/Market-and-Security-Analysis/excel files/put_call.xlsx')
    ws = wb.active

    # Get First Empty Row to Append @
    for row in range(1, 10000):
        cell = ws[f'A{row}']
        if cell.value is None:
            first_empty_row = row
            break

    # today = datetime.date.today()
    time.sleep(0.5)
    total_pc, equity_pc = download_put_call()

    ws[f'A{first_empty_row}'], ws[f'B{first_empty_row}'], ws[f'C{first_empty_row}'] = single_date, total_pc, equity_pc
    wb.save('C:/Users/joema/PycharmProjects/Market-and-Security-Analysis/excel files/put_call.xlsx')


def date_range(start_date, end_date):
    for n in range(int((end_date - start_date).days)):
        yield start_date + timedelta(n)


def plot(timeframe, option_type):
    data = pd.read_excel('excel files/put_call.xlsx', engine='openpyxl')
    dates = pd.to_datetime(data['Day'])
    tpc = data['Total Put/Call']
    epc = data['Equity Put/Call']
    tpc_rolling = tpc.rolling(window=timeframe).mean()
    epc_rolling = epc.rolling(window=timeframe).mean()
    fig, ax = plt.subplots(figsize=(11, 6))

    if option_type == 'tpc':
        ax.plot(dates, tpc, label='Total Put/Call', color='black')
        ax.plot(dates, tpc_rolling, label=f'{timeframe}-Day Moving Average', color='purple')
        plt.title('CBOE Total Put/Call Ratio\n', fontsize=21)
        plt.figtext(0.335, 0.89, f'Current Put/Call Ratio: {list(tpc)[-1]}', fontsize=21, color='purple')

    elif option_type == 'epc':
        ax.plot(dates, epc, label='Equity Put/Call', color='black')
        ax.plot(dates, epc_rolling, label=f'{timeframe}-Day Moving Average', color='purple')
        plt.title('Equity Put/Call Ratio\n', fontsize=21)
        plt.figtext(0.335, 0.89, f'Current Put/Call Ratio: {list(epc)[-1]}', fontsize=21, color='purple')

    ax.xaxis.set_major_locator(matplotlib.dates.DayLocator(interval=date_axis_plotting_interval))
    plt.legend(facecolor='#bebebe', framealpha=0, fontsize=15)
    ax.set_facecolor('#bebebe')
    plt.tight_layout()
    plt.savefig(f'plots/{option_type}_put_call.png')


# /////////////////////////////////////////////////////////////////////////////////////////////////////////////////


if get_historical:
    for single_date in date_range(start_date, end_date):
        # Monday is 0, Sunday is 6
        if single_date.weekday() in [0, 1, 2, 3, 4]:
            update_excel_sheet()
elif get_today:
    single_date = datetime.date.today()
    if single_date.weekday() in [0, 1, 2, 3, 4]:
        update_excel_sheet()
elif get_today and plot_figure:
    single_date = datetime.date.today()
    if single_date.weekday() in [0, 1, 2, 3, 4]:
        update_excel_sheet()
    plot(timeframe_to_plot_sma, which_option_to_plot)
else:
    plot(timeframe_to_plot_sma, which_option_to_plot)


plot(timeframe_to_plot_sma, 'tpc')
plot(timeframe_to_plot_sma, 'epc')
