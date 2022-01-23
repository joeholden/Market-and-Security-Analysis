import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import datetime
from datetime import date, timedelta
import matplotlib.pyplot as plt
import matplotlib


def scrape_wsj_market_breadth():
    url = 'https://www.wsj.com/market-data/stocks/marketsdiary'
    s = Service(ChromeDriverManager().install())
    op = webdriver.ChromeOptions()
    op.headless = True
    driver = webdriver.Chrome(service=s, options=op)
    driver.get(url)

    nyse_advancers = driver.find_element(by=By.XPATH,
                                         value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[1]/tr[2]/td[2]')
    nyse_decliners = driver.find_element(by=By.XPATH,
                                         value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[1]/tr[3]/td[2]')
    nyse_new_highs = driver.find_element(by=By.XPATH,
                                         value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[1]/tr[5]/td[2]')
    nyse_new_lows = driver.find_element(by=By.XPATH,
                                        value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[1]/tr[6]/td[2]')
    nyse_adv_vol = driver.find_element(by=By.XPATH,
                                       value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[1]/tr[7]/td[2]')
    nyse_dec_vol = driver.find_element(by=By.XPATH,
                                       value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[1]/tr[8]/td[2]')

    nasdaq_advancers = driver.find_element(by=By.XPATH,
                                           value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[2]/tr[2]/td[2]')
    nasdaq_decliners = driver.find_element(by=By.XPATH,
                                           value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[2]/tr[3]/td[2]')
    nasdaq_new_highs = driver.find_element(by=By.XPATH,
                                           value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[2]/tr[5]/td[2]')
    nasdaq_new_lows = driver.find_element(by=By.XPATH,
                                          value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[2]/tr[6]/td[2]')
    nasdaq_adv_vol = driver.find_element(by=By.XPATH,
                                         value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[2]/tr[9]/td[2]')
    nasdaq_dec_vol = driver.find_element(by=By.XPATH,
                                         value='//*[@id="root"]/div/div/div/div[2]/div/div/div[2]/table/tbody[2]/tr[10]/td[2]')

    return (int(nyse_advancers.text.replace(',', '')), int(nyse_decliners.text.replace(',', '')),
            int(nyse_adv_vol.text.replace(',', '')), int(nyse_dec_vol.text.replace(',', '')),
            int(nyse_new_highs.text.replace(',', '')), int(nyse_new_lows.text.replace(',', '')), \
            int(nasdaq_advancers.text.replace(',', '')), int(nasdaq_decliners.text.replace(',', '')),
            int(nasdaq_adv_vol.text.replace(',', '')), int(nasdaq_dec_vol.text.replace(',', '')),
            int(nasdaq_new_highs.text.replace(',', '')), int(nasdaq_new_lows.text.replace(',', '')))


print(scrape_wsj_market_breadth())


def update_excel_sheet():
    headers = ['NYSE Advancers', 'NYSE Decliners', 'NYSE Adv. Volume', 'NYSE Dec. Volume', 'NYSE New Highs',
               'NYSE New Lows', 'NASDAQ Advancers', 'NASDAQ Decliners', 'NASDAQ Adv. Volume', 'NASDAQ Dec. Volume',
               'NASDAQ New Highs', 'NASDAQ New Lows']

    wb = load_workbook('C:/Users/joema/PycharmProjects/stocks/excel files/WSJ Breadth.xlsx')
    ws = wb.active

    # Get First Empty Row to Append @
    for row in range(1, 10000):
        cell = ws[f'A{row}']
        if cell.value is None:
            first_empty_row = row
            break

    data = scrape_wsj_market_breadth()
    ws[f'A{first_empty_row}'] = datetime.date.today()

    ws[f'B{first_empty_row}'], ws[f'C{first_empty_row}'], ws[f'D{first_empty_row}'], \
    ws[f'E{first_empty_row}'], ws[f'F{first_empty_row}'], ws[f'G{first_empty_row}'], \
    ws[f'H{first_empty_row}'], ws[f'I{first_empty_row}'], ws[f'J{first_empty_row}'], \
    ws[f'K{first_empty_row}'], ws[f'L{first_empty_row}'], ws[f'M{first_empty_row}'] = data

    wb.save('C:/Users/joema/PycharmProjects/stocks/excel files/WSJ Breadth.xlsx')


def plot():
    pass
